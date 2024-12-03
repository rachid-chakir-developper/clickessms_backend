import graphene
from graphene_django import DjangoObjectType
from graphql_jwt.decorators import login_required
from graphene_file_upload.scalars import Upload
import re
from django.utils.dateparse import parse_date
from datetime import datetime

from django.apps import apps

from django.db.models import Q
from django.db import transaction
import openpyxl

from finance.models import BudgetAccountingNature
from data_management.models import HumanGender, AdmissionDocumentType, PhoneNumber, HomeAddress, DataModel, EstablishmentType, EstablishmentCategory, AbsenceReason, UndesirableEventNormalType, UndesirableEventSeriousType, UndesirableEventFrequency, MeetingReason, TypeMeeting, DocumentType, VehicleBrand, VehicleModel, EmployeeMission, AccountingNature, CustomField, CustomFieldOption, CustomFieldValue



class HumanGenderType(DjangoObjectType):
    class Meta:
        model = HumanGender
        fields = "__all__"

class AdmissionDocumentTypeType(DjangoObjectType):
    class Meta:
        model = AdmissionDocumentType
        fields = "__all__"
        
class PhoneNumberType(DjangoObjectType):
    class Meta:
        model = PhoneNumber
        fields = "__all__"
    name = graphene.String()
    def resolve_name( instance, info, **kwargs ):
        return instance.name if instance.name else 'Sans nom'

class PhoneNumberNodeType(graphene.ObjectType):
    nodes = graphene.List(PhoneNumberType)
    total_count = graphene.Int()
        
class HomeAddressType(DjangoObjectType):
    class Meta:
        model = HomeAddress
        fields = "__all__"

class HomeAddressNodeType(graphene.ObjectType):
    nodes = graphene.List(HomeAddressType)
    total_count = graphene.Int()

class DataType(DjangoObjectType):
    class Meta:
        model = DataModel
        fields = "__all__"

class AbsenceReasonType(DjangoObjectType):
    class Meta:
        model = AbsenceReason
        fields = "__all__"
        
class EstablishmentCategoryType(DjangoObjectType):
    class Meta:
        model = EstablishmentCategory
        fields = "__all__"

class EstablishmentTypeType(DjangoObjectType):
    class Meta:
        model = EstablishmentType
        fields = "__all__"

class UndesirableEventNormalTypeType(DjangoObjectType):
    class Meta:
        model = UndesirableEventNormalType
        fields = "__all__"

class UndesirableEventSeriousTypeType(DjangoObjectType):
    class Meta:
        model = UndesirableEventSeriousType
        fields = "__all__"

class UndesirableEventFrequencyType(DjangoObjectType):
    class Meta:
        model = UndesirableEventFrequency
        fields = "__all__"

class MeetingReasonType(DjangoObjectType):
    class Meta:
        model = MeetingReason
        fields = "__all__"

class TypeMeetingType(DjangoObjectType):
    class Meta:
        model = TypeMeeting
        fields = "__all__"
class DocumentTypeType(DjangoObjectType):
    class Meta:
        model = DocumentType
        fields = "__all__"
        
class VehicleBrandType(DjangoObjectType):
    class Meta:
        model = VehicleBrand
        fields = "__all__"
     
class VehicleModelType(DjangoObjectType):
    class Meta:
        model = VehicleModel
        fields = "__all__"
     
class EmployeeMissionType(DjangoObjectType):
    class Meta:
        model = EmployeeMission
        fields = "__all__"
     
class AccountingNatureType(DjangoObjectType):
    class Meta:
        model = AccountingNature
        fields = "__all__"
    children_number = graphene.Int()
    amount_allocated = graphene.Decimal(required=False)
    def resolve_children_number( instance, info, **kwargs ):
        return instance.children.count() if hasattr(instance, "children") else 0
    def resolve_amount_allocated(instance, info, **kwargs):
        accounting_nature_filter = getattr(info.context, 'accounting_nature_filter', None)
        if accounting_nature_filter:
            budget_id = accounting_nature_filter.get('budget', None)
            if budget_id:
                try:
                    budget_accounting_nature = BudgetAccountingNature.objects.filter(
                        budget_id=budget_id,
                        accounting_nature=instance
                    ).first()
                    return budget_accounting_nature.amount_allocated if budget_accounting_nature else None
                except BudgetAccountingNature.DoesNotExist:
                    return None
        return None

class AccountingNatureFilterInput(graphene.InputObjectType):
        keyword = graphene.String(required=False)
        list_type = graphene.String(required=False)
        budget = graphene.Int(required=False)

class AccountingNatureNodeType(graphene.ObjectType):
    nodes = graphene.List(AccountingNatureType)
    total_count = graphene.Int()

class CustomFieldOptionType(DjangoObjectType):
    class Meta:
        model = CustomFieldOption
        fields = "__all__"

class CustomFieldType(DjangoObjectType):
    class Meta:
        model = CustomField
        fields = "__all__"
    form_model = graphene.String()
    def resolve_form_model( instance, info, **kwargs ):
        return instance.form_model

class CustomFieldValueType(DjangoObjectType):
    class Meta:
        model = CustomFieldValue
        fields = "__all__"
    key = graphene.String()
    def resolve_key( instance, info, **kwargs ):
        return instance.key

class CustomFieldValueInput(graphene.InputObjectType):
    id = graphene.ID(required=False)
    key = graphene.String(required=False)
    value = graphene.String(required=False)
    custom_field_id = graphene.Int(name="customField", required=False)

class CustomFieldFilterInput(graphene.InputObjectType):
    keyword = graphene.String(required=False)
    form_models = graphene.List(graphene.String, required=False)

class CustomFieldNodeType(graphene.ObjectType):
    nodes = graphene.List(CustomFieldType)
    total_count = graphene.Int()

class CustomFieldOptionInput(graphene.InputObjectType):
    id = graphene.ID(required=False)
    label = graphene.String(required=True)
    value = graphene.String(required=False)

class CustomFieldInput(graphene.InputObjectType):
    id = graphene.ID(required=False)
    label = graphene.String(required=True)
    key = graphene.String(required=False)
    field_type = graphene.String(required=True)
    form_model = graphene.String(required=True)
    is_active = graphene.Boolean(required=False)
    options = graphene.List(CustomFieldOptionInput, required=False)

class AccountingNatureInput(graphene.InputObjectType):
    id = graphene.ID(required=False)
    code = graphene.String(required=True)
    name = graphene.String(required=True)
    description = graphene.String(required=True)
    is_active = graphene.Boolean(required=False)
    parent_id = graphene.Int(name="parent", required=False)

class DataQuery(graphene.ObjectType):
    datas = graphene.List(DataType, typeData = graphene.String(), id_parent = graphene.ID(required=False))
    data = graphene.Field(DataType, typeData = graphene.String(), id = graphene.ID())
    accounting_natures = graphene.Field(AccountingNatureNodeType, accounting_nature_filter= AccountingNatureFilterInput(required=False), id_parent = graphene.ID(required=False), offset = graphene.Int(required=False), limit = graphene.Int(required=False), page = graphene.Int(required=False))
    accounting_nature = graphene.Field(AccountingNatureType, id = graphene.ID())
    custom_fields = graphene.Field(CustomFieldNodeType, custom_field_filter= CustomFieldFilterInput(required=False), id_company = graphene.ID(required=False), offset = graphene.Int(required=False), limit = graphene.Int(required=False), page = graphene.Int(required=False))
    custom_field = graphene.Field(CustomFieldType, id = graphene.ID())
    
    def resolve_datas(root, info, typeData, id_parent=None):
        user = info.context.user
        company = user.the_current_company
        # We can easily optimize query count in the resolve method
        datas = apps.get_model('data_management', typeData).objects.filter(Q(company=company) | Q(creator__is_superuser=True))
        if id_parent:
            datas = datas.filter(parent_id=id_parent)
        else:
            datas = datas.filter(parent__isnull=True)
        for data in datas:
            parent = data.parent
            data.__class__ = DataModel
        return datas

    def resolve_data(root, info, typeData, id):
        # We can easily optimize query count in the resolve method
        data = apps.get_model('data_management', typeData).objects.get(pk=id)
        data.__class__ = DataModel
        return data
    
    def resolve_accounting_natures(root, info, accounting_nature_filter=None, id_parent=None, offset=None, limit=None, page=None):
        # We can easily optimize query count in the resolve method
        user = info.context.user
        info.context.accounting_nature_filter = accounting_nature_filter
        company = user.the_current_company
        total_count = 0
        accounting_natures = AccountingNature.objects.filter(Q(company=company) | Q(creator__is_superuser=True))
        if id_parent:
            accounting_natures = accounting_natures.filter(parent_id=id_parent)
        else:
            accounting_natures = accounting_natures.filter(parent__isnull=True)
        if accounting_nature_filter:
            keyword = accounting_nature_filter.get('keyword', '')
            list_type = accounting_nature_filter.get('list_type', None)
            if list_type and list_type=='ALL':
                accounting_natures = AccountingNature.objects.filter(Q(company=company) | Q(creator__is_superuser=True), is_active=True)
            if keyword:
                accounting_natures = accounting_natures.filter(Q(code__icontains=keyword) | Q(name__icontains=keyword) | Q(description__icontains=keyword))
        accounting_natures = accounting_natures.order_by('code').distinct()
        total_count = accounting_natures.count()
        if page:
            offset = limit * (page - 1)
        if offset is not None and limit is not None:
            accounting_natures = accounting_natures[offset:offset + limit]
        return AccountingNatureNodeType(nodes=accounting_natures, total_count=total_count)

    def resolve_accounting_nature(root, info, id):
        # We can easily optimize query count in the resolve method
        try:
            accounting_nature = AccountingNature.objects.get(pk=id)
        except AccountingNature.DoesNotExist:
            accounting_nature = None
        return accounting_nature

    def resolve_custom_fields(root, info, custom_field_filter=None, id_company=None, offset=None, limit=None, page=None):
        # We can easily optimize query count in the resolve method
        user = info.context.user
        company = user.the_current_company
        total_count = 0
        custom_fields = CustomField.objects.filter(company__id=id_company) if id_company else CustomField.objects.filter(company=company)
        if custom_field_filter:
            keyword = custom_field_filter.get('keyword', '')
            form_models = custom_field_filter.get('form_models')
            if form_models:
                custom_fields = custom_fields.filter(form_model__in=form_models)
            if keyword:
                custom_fields = custom_fields.filter(Q(title__icontains=keyword))
        custom_fields = custom_fields.order_by('created_at')
        total_count = custom_fields.count()
        if page:
            offset = limit * (page - 1)
        if offset is not None and limit is not None:
            custom_fields = custom_fields[offset:offset + limit]
        return CustomFieldNodeType(nodes=custom_fields, total_count=total_count)

    def resolve_custom_field(root, info, id):
        # We can easily optimize query count in the resolve method
        try:
            custom_field = CustomField.objects.get(pk=id)
        except CustomField.DoesNotExist:
            custom_field = None
        return custom_field

class CreateData(graphene.Mutation):
    class Arguments:
        name = graphene.String()
        description = graphene.String(required=False)
        code = graphene.String(required=False)
        parent_id = graphene.ID(required=False)
        typeData = graphene.String()

    data = graphene.Field(DataType)

    def mutate(root, info, typeData, parent_id=None, **otherFields):
        creator = info.context.user
        data = apps.get_model('data_management', typeData)(**otherFields)
        data.creator = creator
        data.company = creator.the_current_company
        data.save()
        data.__class__ = DataModel
        return CreateData(data=data)

class UpdateData(graphene.Mutation):
    class Arguments:
        name = graphene.String()
        description = graphene.String(required=False)
        code = graphene.String(required=False)
        typeData = graphene.String()
        id = graphene.ID()

    data = graphene.Field(DataType)

    def mutate(root, info, id, typeData, **otherFields):
        apps.get_model('data_management', typeData).objects.filter(pk=id).update(**otherFields)
        data = apps.get_model('data_management', typeData).objects.get(pk=id)
        # data.name = name
        # data.description = description
        data.__class__ = DataModel
        return UpdateData(data=data)

class DeleteData(graphene.Mutation):
    class Arguments:
        typeData = graphene.String()
        id = graphene.ID()

    data = graphene.Field(DataType)
    deleted = graphene.Boolean()

    def mutate(root, info, id, typeData):
        data = apps.get_model('data_management', typeData).objects.get(pk=id)
        data.delete()
        return DeleteData(deleted=True)

#**************************************************************************************************

#**************************************************************************************************

class CreateAccountingNature(graphene.Mutation):
    class Arguments:
        accounting_nature_data = AccountingNatureInput(required=True)

    accounting_nature = graphene.Field(AccountingNatureType)

    def mutate(root, info, accounting_nature_data=None):
        creator = info.context.user
        if not creator.is_superuser:
            raise ValueError("Impossible d'ajouter : vous n'avez pas les droits nécessaires.")
        accounting_nature = AccountingNature(**accounting_nature_data)
        accounting_nature.creator = creator
        accounting_nature.company = creator.the_current_company
        accounting_nature.save()
        return CreateAccountingNature(accounting_nature=accounting_nature)

class UpdateAccountingNature(graphene.Mutation):
    class Arguments:
        id = graphene.ID()
        accounting_nature_data = AccountingNatureInput(required=True)

    accounting_nature = graphene.Field(AccountingNatureType)

    def mutate(root, info, id, accounting_nature_data=None):
        creator = info.context.user
        if not creator.is_superuser:
            raise ValueError("Impossible de modifier : vous n'avez pas les droits nécessaires.")
        accounting_nature_data.pop('parent_id')
        AccountingNature.objects.filter(pk=id).update(**accounting_nature_data)
        accounting_nature = AccountingNature.objects.get(pk=id)
        return UpdateAccountingNature(accounting_nature=accounting_nature)

class DeleteAccountingNature(graphene.Mutation):
    class Arguments:
        id = graphene.ID()

    accounting_nature = graphene.Field(AccountingNatureType)
    id = graphene.ID()
    deleted = graphene.Boolean()
    success = graphene.Boolean()
    message = graphene.String()

    def mutate(root, info, id):
        deleted = False
        success = False
        message = ''
        current_user = info.context.user
        if current_user.is_superuser:
            accounting_nature = AccountingNature.objects.get(pk=id)
            accounting_nature.delete()
            deleted = True
            success = True
        else:
            message = "Impossible de supprimer : vous n'avez pas les droits nécessaires."
        return DeleteAccountingNature(deleted=deleted, success=success, message=message, id=id)

#**************************************************************************************************


def extract_parts_name(full_name):
    words = full_name.split()
    first_name = None
    preferred_name = None
    last_name = None
    for i, word in enumerate(words):
        if word.lower() in ["né", "née"]:
            last_name = " ".join(words[i+1:]).upper()
            break
        elif word.isupper():
            preferred_name = word
        elif word[0].isupper() and i > 0:
            first_name = word
    if not preferred_name and not first_name:
        first_name = words[0].capitalize() if words else None
        preferred_name = words[1].upper() if len(words) > 1 else None
    return first_name, preferred_name, last_name

def import_data_from_file(entity, model, file, fields, user=None):
    count = 0
    new_objects = []
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    with transaction.atomic():
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            data = {field: row_data[field] for field in fields if field in row_data}
            # model.objects.create(**data)
            if entity == 'Employee':
                try:
                    registration_number, name, social_security_number = data['registration_number'], data['name'], data['social_security_number']
                    first_name, last_name, preferred_name = extract_parts_name(full_name=name)
                    employee = model.objects.get(Q(first_name=first_name, last_name=last_name))
                    if not employee.social_security_number or employee.social_security_number == '' :
                        model.objects.filter(pk=employee.id).update(social_security_number=social_security_number)
                    if not employee.registration_number or employee.registration_number == '' :
                        model.objects.filter(pk=employee.id).update(registration_number=registration_number)
                except model.DoesNotExist:
                    # new_objects.append(model(
                    #     registration_number=registration_number,
                    #     first_name=first_name,
                    #     last_name=preferred_name if preferred_name else last_name,
                    #     preferred_name=last_name if preferred_name else preferred_name,
                    #     social_security_number=social_security_number,
                    #     company=user.company,
                    #     creator=user,
                    #     ))
                    pass
                except Exception as e:
                    pass
                count += 1
            if entity == 'Beneficiary':
                try:
                    name, address, zip_code, city, birth_date = data['name'], data['address'], data['zip_code'], data['city'], data['birth_date']
                    # birth_date = parse_date(birth_date)
                    # first_name, last_name, preferred_name = extract_parts_name(full_name=name)
                    first_name, last_name, preferred_name = None, None, None
                    if name and name != '':
                        try:
                            birth_date = datetime.strptime(birth_date, '%d/%m/%Y').date() if birth_date else None
                        except Exception as e:
                            birth_date = None
                        full_name = name
                        words = full_name.split()
                        first_name = words[1]
                        last_name = words[0]
                        preferred_name = last_name
                    beneficiary = model.objects.get(Q(first_name=first_name, last_name=last_name))
                    if not beneficiary.birth_date:
                        model.objects.filter(pk=beneficiary.id).update(birth_date=birth_date)
                except model.DoesNotExist:
                    if first_name and first_name:
                        new_objects.append(model(
                            first_name=first_name,
                            last_name=last_name,
                            preferred_name=last_name,
                            birth_date=birth_date,
                            address=address,
                            zip_code=zip_code,
                            city=city,
                            company=user.company,
                            creator=user,
                            ))
                        count += 1
                except Exception as e:
                    raise e
        model.objects.bulk_create(new_objects)
    return count

class ImportDataMutation(graphene.Mutation):
    class Arguments:
        entity = graphene.String(required=True)
        file = Upload(required=True)
        fields = graphene.List(graphene.String, required=True)

    count = graphene.Int()
    done = graphene.Boolean()
    success = graphene.Boolean()

    def mutate(self, info, entity=None, file=None, fields=None):
        user = info.context.user
        done = True
        success = True
        count = 0
        try:
            if entity == 'Employee' or entity == 'Beneficiary':
                model_app = 'human_ressources'
            else:
                done = False
                success = False
            model = apps.get_model(model_app, entity)
            count = import_data_from_file(entity=entity, model=model, file=file, fields=fields, user=user)
        except Exception as e:
            print(e)
            done = False
            success = False
        return ImportDataMutation(success=success, done=done, count=count)

#**************************************************************************************************

class CreateCustomField(graphene.Mutation):
    class Arguments:
        custom_field_data = CustomFieldInput(required=True)

    custom_field = graphene.Field(CustomFieldType)

    def mutate(root, info, custom_field_data=None):
        creator = info.context.user
        options = custom_field_data.pop("options")
        custom_field = CustomField(**custom_field_data)
        custom_field.creator = creator
        custom_field.company = creator.the_current_company
        custom_field.save()
        for option in options:
            custom_field_option = CustomFieldOption(**option)
            custom_field_option.custom_field = custom_field
            custom_field_option.save()
        return CreateCustomField(custom_field=custom_field)

class UpdateCustomField(graphene.Mutation):
    class Arguments:
        id = graphene.ID()
        custom_field_data = CustomFieldInput(required=True)
        image = Upload(required=False)

    custom_field = graphene.Field(CustomFieldType)

    def mutate(root, info, id, image=None, custom_field_data=None):
        creator = info.context.user
        options = custom_field_data.pop("options")
        CustomField.objects.filter(pk=id).update(**custom_field_data)
        custom_field = CustomField.objects.get(pk=id)
        option_ids = [
            item.id for item in options if item.id is not None
        ]
        CustomFieldOption.objects.filter(
            custom_field=custom_field
        ).exclude(id__in=option_ids).delete()
        for option in options:
            if id in option or "id" in option:
                CustomFieldOption.objects.filter(pk=option.id).update(**option)
            else:
                custom_field_option = CustomFieldOption(**option)
                custom_field_option.custom_field = custom_field
                custom_field_option.save()
        return UpdateCustomField(custom_field=custom_field)

class DeleteCustomField(graphene.Mutation):
    class Arguments:
        id = graphene.ID()

    custom_field = graphene.Field(CustomFieldType)
    id = graphene.ID()
    deleted = graphene.Boolean()
    success = graphene.Boolean()
    message = graphene.String()

    def mutate(root, info, id):
        deleted = False
        success = False
        message = ''
        current_user = info.context.user
        if current_user.is_superuser:
            custom_field = CustomField.objects.get(pk=id)
            custom_field.delete()
            deleted = True
            success = True
        else:
            message = "Impossible de supprimer : vous n'avez pas les droits nécessaires."
        return DeleteCustomField(deleted=deleted, success=success, message=message, id=id)

#**************************************************************************************************

class DataMutation(graphene.ObjectType):
    create_data = CreateData.Field()
    update_data = UpdateData.Field()
    delete_data = DeleteData.Field()

    create_accounting_nature = CreateAccountingNature.Field()
    update_accounting_nature = UpdateAccountingNature.Field()
    delete_accounting_nature = DeleteAccountingNature.Field()

    import_data = ImportDataMutation.Field()

    create_custom_field = CreateCustomField.Field()
    update_custom_field = UpdateCustomField.Field()
    delete_custom_field = DeleteCustomField.Field()