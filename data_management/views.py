from django.http import HttpResponse
from django.apps import apps
from django.views import View
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
import json
from django.db import models  # Import the models module


class ExportDataView(View):
    def post(self, request):
        if not request.user.is_authenticated:
            return HttpResponse("Utilisateur non authentifié", status=401)
        
        # Parse the incoming JSON request
        user = request.user
        company = user.get_current_company()
        body = json.loads(request.body)
        entity_name = body.get('entity')
        file_name = body.get('file_name')
        fields = body.get('fields')
        titles = body.get('titles', [])  # Optional titles, defaults to empty list

        if not entity_name or not fields:
            return HttpResponse("Les champs 'entity' et 'fields' sont requis.", status=400)

        # Try to get the model from the entity name
        model = None
        for app in apps.get_app_configs():
            try:
                model = apps.get_model(app.label, entity_name)
                break
            except LookupError:
                continue

        if not model:
            return HttpResponse("Modèle introuvable", status=404)

        # Retrieve all data for the specified model
        queryset = model.objects.filter(company=company, is_deleted=False)

        # Create an in-memory Excel workbook and sheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f'{entity_name} Data'

        # Write column headers
        for col_num, (field, title) in enumerate(zip(fields, titles), 1):
            column_letter = get_column_letter(col_num)
            worksheet[f'{column_letter}1'] = title if title else field

        # Function to extract the value of a field, including nested fields
        def get_field_value(obj, field):
            if isinstance(field, list):  # Handle list of fields, e.g., [first_name, last_name]
                values = []
                for f in field:
                    if '__' in f:
                        related_parts = f.split('__')
                        related_obj = getattr(obj, related_parts[0], None)
                        if related_obj is not None:
                            if isinstance(related_obj, (list, models.Manager)):  # For ManyToMany or reverse FK
                                for item in related_obj.all():
                                    sub_values = get_field_value(item, '__'.join(related_parts[1:]))
                                    values.append(sub_values)
                            else:
                                sub_values = get_field_value(related_obj, '__'.join(related_parts[1:]))
                                values.append(sub_values)
                    else:
                        if hasattr(obj, f'get_{f}_display'):
                            values.append(getattr(obj, f'get_{f}_display')())
                        # Single field case
                        else:
                            value = getattr(obj, f, '')
                            if isinstance(value, datetime):
                                # Remove timezone info if present
                                value = value.replace(tzinfo=None)
                            values.append(value)
                return '; '.join(map(str, values))  # Combine multiple field values
            else:
                parts = field.split('__')  # Split nested fields (e.g., vehicle_employees__employees__first_name)
                value = obj
                for part in parts:
                    if value is None:
                        return ''
                    if hasattr(value, 'all'):  # Handle ManyToMany relationships
                        related_values = []
                        for related_obj in value.all():
                            related_values.append(get_field_value(related_obj, '__'.join(parts[1:])))
                        return '| '.join(related_values)
                    else:
                        if hasattr(value, f'get_{part}_display'):
                            value = getattr(value, f'get_{part}_display')()
                        else:
                            value = getattr(value, part, None)
                            if isinstance(value, datetime):
                                # Remove timezone info if present
                                value = value.replace(tzinfo=None)
                return value or ''

        # Write the data rows
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field in enumerate(fields, 1):
                column_letter = get_column_letter(col_num)
                worksheet[f'{column_letter}{row_num}'] = get_field_value(obj, field)

        # Create the HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name if file_name else entity_name}_data.xlsx"'

        # Save the workbook to the response
        workbook.save(response)
        return response
