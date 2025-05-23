
import base64
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import datetime
from io import BytesIO

from django.core.exceptions import PermissionDenied

from django.db.models import Q, Case, When
from dashboard.utils import get_age_range
from calendar import monthrange
from django.conf import settings

from companies.models import Establishment
from human_ressources.models import BeneficiaryEntry, BeneficiaryAdmission


def generate_excel_activity_month(info=None, dashboard_activity_filter=None, data=None):
    user = info.context.user
    company = user.the_current_company
    if not user.can_manage_activity() and not user.is_manager():
        raise PermissionDenied("Impossible d'exporter : vous n'avez pas les droits nécessaires.")
    try:
        wb = Workbook()
        # Supprimer la feuille par défaut
        wb.remove(wb.active)

        # Obtenir l'année en cours pour filtrer par année
        date = datetime.date.today()
        year=str(date.year)
        month=date.month - 1
        if month <= 0:  
            month = 12  
            year -= 1  
        elif month > 12:  
            month = 1  
            year += 1 
        start_year = date.replace(month=1, day=1)  # Début de l'année
        end_year = date.replace(month=12, day=31)  # Fin de l'année

        establishments = Establishment.objects.filter(company=company, establishment_parent__id=None, is_deleted=False)
        if user.is_manager() and not user.can_manage_activity():
            establishments = establishments.filter(managers__employee=user.get_employee_in_company())
        if dashboard_activity_filter:
            the_year = dashboard_activity_filter.get('year', None)
            the_month = dashboard_activity_filter.get('month', None)
            establishment_ids = dashboard_activity_filter.get('establishments', None)
            if the_year:
                year=the_year
            if the_month:
                month=the_month
            # if establishment_ids:
            #     order = Case(*[When(id=pk, then=pos) for pos, pk in enumerate(establishment_ids)])
            #     establishments=establishments.filter(id__in=establishment_ids).annotate(ordering=order).order_by('ordering')

        present_beneficiaries = BeneficiaryEntry.present_beneficiaries(year=year, month=month, establishments=establishments, company=company)
        month_beneficiary_admissions = BeneficiaryAdmission.month_beneficiary_admissions(year=year, month=month, establishments=establishments, company=company, statuses=["APPROVED"])
        
        for i, establishment in enumerate(establishments):
            beneficiary_admissions = []
            admissions_waiting_data = []
            children_establishments = establishment.get_all_children()
            beneficiary_entries=present_beneficiaries.get(establishment.id, [])
            beneficiary_admissions=month_beneficiary_admissions.get(establishment.id, [])
            for i, beneficiary_admission in enumerate(beneficiary_admissions):
                beneficiary = beneficiary_admission.beneficiary
                admissions_waiting_data.append(
                    [
                        establishment.name,
                        beneficiary_admission.last_name,
                        beneficiary_admission.first_name,  # Correction ici
                        beneficiary_admission.birth_date and beneficiary_admission.birth_date.date().strftime('%d/%m/%Y'),
                        f"{beneficiary.get_custom_field_value('IEF')}" if beneficiary else "",
                        beneficiary_admission.pre_admission_date and beneficiary_admission.pre_admission_date.date().strftime('%d/%m/%Y'),
                        beneficiary_admission.response_date and beneficiary_admission.response_date.date().strftime('%d/%m/%Y'),
                    ]
                )
            capacity = establishment.get_monthly_capacity(year, month)
            count_occupied_places= len(beneficiary_entries)
            count_available_places=capacity-count_occupied_places

            last_day = monthrange(int(year), int(month))[1]

            # Créer une nouvelle feuille pour l’établissement
            sheet_title = establishment.name[:31]  # Excel limite les noms à 31 caractères
            ws = wb.create_sheet(title=sheet_title)

            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Mise à jour date
            ws["A1"] = f"MAJ"
            ws["B1"] = f"{datetime.date.today().strftime('%d/%m/%Y')}"

            # EFFECTIFS AU...
            ws["A3"] = f"EFFECTIFS AU {str(last_day)}/{int(month):02d}/{year}"
            ws.merge_cells("A3:C3")
            ws["A3"].font = Font(bold=True, italic=True)

            # Tableau de capacité
            ws["B5"] = "CAPACITÉ"
            ws["B5"].border = border

            ws["C5"] = "Places occupées"
            ws["C5"].border = border

            ws["D5"] = "Capacité Totale"
            ws["D5"].border = border

            # Liste des jeunes (données d'exemple)
            start_row = 6

            if len(children_establishments) < 1:
                ws["B6"] = f"{establishment.name}"
                ws["B6"].border = border

                ws["C6"] = count_occupied_places
                ws["D6"] = capacity
                ws["D6"].font = Font(color="FF0000")
                ws["D6"].border = border
                start_row+=1
            else:
                beneficiary_admissions_children = BeneficiaryAdmission.month_beneficiary_admissions(year=year, month=month, establishments=children_establishments, company=company, statuses=["APPROVED"])

            present_beneficiaries_children = BeneficiaryEntry.present_beneficiaries(year=year, month=month, establishments=children_establishments, company=company)
            for i, children_establishment in enumerate(children_establishments):
                children_beneficiary_entries=present_beneficiaries_children.get(children_establishment.id, [])
                beneficiary_admissions=beneficiary_admissions_children.get(children_establishment.id, [])
                for i, beneficiary_admission in enumerate(beneficiary_admissions):
                    beneficiary = beneficiary_admission.beneficiary
                    admissions_waiting_data.append(
                        [
                            children_establishment.name,
                            beneficiary_admission.last_name,
                            beneficiary_admission.first_name,  # Correction ici
                            beneficiary_admission.birth_date and beneficiary_admission.birth_date.date().strftime('%d/%m/%Y'),
                            f"{beneficiary.get_custom_field_value('IEF')}" if beneficiary else "",
                            beneficiary_admission.pre_admission_date and beneficiary_admission.pre_admission_date.date().strftime('%d/%m/%Y'),
                            beneficiary_admission.response_date and beneficiary_admission.response_date.date().strftime('%d/%m/%Y'),
                        ]
                    )
                children_capacity = children_establishment.get_monthly_capacity(year, month)
                children_count_occupied_places= len(children_beneficiary_entries)
                children_count_available_places=children_capacity-children_count_occupied_places
                ws[f"B{start_row}"] = f"{children_establishment.name}"
                ws[f"B{start_row}"].border = border

                ws[f"C{start_row}"] = children_count_occupied_places
                ws[f"C{start_row}"].alignment = Alignment(horizontal="center")
                ws[f"C{start_row}"].border = border
                ws[f"D{start_row}"] = children_capacity
                ws[f"D{start_row}"].font = Font(color="FF0000")
                ws[f"D{start_row}"].alignment = Alignment(horizontal="center")
                ws[f"D{start_row}"].border = border
                start_row+=1

            ws[f"B{start_row}"] = "Place disponible"
            ws[f"B{start_row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            ws[f"B{start_row}"].border = border

            ws[f"C{start_row}"] = count_available_places
            ws.merge_cells(f"C{start_row}:D{start_row}")
            ws[f"C{start_row}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            ws[f"C{start_row}"].alignment = Alignment(horizontal="center")
            ws[f"C{start_row}"].border = border
            ws[f"D{start_row}"].border = border

            # Titre principal
            ws["D1"] = f"{establishment.name}"
            ws.merge_cells("D1:M1")
            ws["D1"].alignment = Alignment(horizontal="center")
            ws["D1"].font = Font(bold=True, size=14)
            ws["D1"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            for col in range(4, 14):
                cell = ws.cell(row=1, column=col)
                cell.border = border

            # Titre de la liste d'attente
            ws["G4"] = "LISTE D'ATTENTE"
            ws.merge_cells("G4:M4")
            ws["G4"].alignment = Alignment(horizontal="center")
            ws["G4"].font = Font(bold=True)
            for col in range(7, 14):
                cell = ws.cell(row=4, column=col)
                cell.border = border

            # En-tête liste d'attente
            headers_waiting = ["UNITE", "NOM", "PRENOM", "DATE DE NAISSANCE", "IEF", "RDV PRÉ-ADMISSION", "Date Admission"]
            for i, header in enumerate(headers_waiting, start=7):
                cell = ws.cell(row=5, column=i, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="e1e1e1", end_color="e1e1e1", fill_type="solid")
                cell.border = border

            ws["G6"] = "Merci d'indiquer les dossiers en attente (préad, admission ou bien en étude)"
            ws.merge_cells("G6:M6")
            ws["G6"].alignment = Alignment(horizontal="center")
            ws["G6"].font = Font(color="FF0000")
            ws["G6"].border = border
            ws["M6"].border = border
            for col in range(7, 14):
                cell = ws.cell(row=6, column=col)
                cell.border = border
            
            for i, row in enumerate(admissions_waiting_data, start=7):
                for j, value in enumerate(row, start=7):
                    cell = ws.cell(row=i, column=j, value=value)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
            
            start_row+= len(admissions_waiting_data)
            start_row+= 3
            cells_to_ignore_width = ["A3", "G6"]
            if len(children_establishments)>0:
                ws[f"A{start_row}"] = f"{establishment.name}"
                cells_to_ignore_width.append(f"A{start_row}")
                ws[f"A{start_row}"].font = Font(bold=True)
                ws[f"A{start_row}"].fill = PatternFill(start_color="e1e1e1", end_color="e1e1e1", fill_type="solid")
                ws[f"A{start_row}"].border = border
                ws.merge_cells(f"A{start_row}:M{start_row}")
                start_row+=1

            headers_main = [
                "Nbre de place", "ETABLISSEMENT", "N°", "UNITE", "NOM", "PRENOM",
                "DATE DE NAISSANCE", "AGE", "ATJM", "IEF",
                "DATE DEBUT ACCUEIL", "DATE DE SORTIE PREVUE", "COMMENTAIRES"
            ]
            for col, header in enumerate(headers_main, start=1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="e1e1e1", end_color="e1e1e1", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Exemple de données
            main_data = []
            for i, entry in enumerate(beneficiary_entries):
                beneficiary = entry.beneficiary
                main_data.append(
                    [
                        i + 1,
                        establishment.name,
                        1,  # Peut-être à remplacer ?
                        establishment.name,
                        beneficiary.last_name,
                        beneficiary.first_name,  # Correction ici
                        beneficiary.birth_date and beneficiary.birth_date.date().strftime('%d/%m/%Y'),
                        int(beneficiary.age_exact),
                        "x" if int(beneficiary.age_exact) >= 18 else "",
                        f"{beneficiary.get_custom_field_value('IEF')}",
                        entry.entry_date and entry.entry_date.date().strftime('%d/%m/%Y'),
                        entry.due_date and entry.due_date.date().strftime('%d/%m/%Y'),
                        f"Sort le {entry.release_date and entry.release_date.date().strftime('%d/%m/%Y')}" if (entry.release_date and False) else ""
                    ]
                )
            next_row = start_row + 1
            for i, row in enumerate(main_data, start=start_row + 1):
                for j, value in enumerate(row, start=1):
                    cell = ws.cell(row=i, column=j, value=value)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                next_row+=1
            for i, children_establishment in enumerate(children_establishments):
                beneficiary_entries=present_beneficiaries_children.get(children_establishment.id, [])
                ws[f"A{next_row}"] = f"{children_establishment.name}"
                cells_to_ignore_width.append(f"A{next_row}")
                ws[f"A{next_row}"].font = Font(bold=True)
                ws[f"A{next_row}"].fill = PatternFill(start_color="e1e1e1", end_color="e1e1e1", fill_type="solid")
                ws[f"A{next_row}"].border = border
                ws.merge_cells(f"A{next_row}:M{next_row}")
                main_data = []
                for i, entry in enumerate(beneficiary_entries):
                    beneficiary = entry.beneficiary
                    main_data.append(
                        [
                            i + 1,
                            establishment.name,
                            1,  # Peut-être à remplacer ?
                            children_establishment.name,
                            beneficiary.last_name,
                            beneficiary.first_name,  # Correction ici
                            beneficiary.birth_date and beneficiary.birth_date.date().strftime('%d/%m/%Y'),
                            int(beneficiary.age_exact),
                            "x" if int(beneficiary.age_exact) >= 18 else "",
                            f"{beneficiary.get_custom_field_value('IEF')}",
                            entry.entry_date and entry.entry_date.date().strftime('%d/%m/%Y'),
                            entry.due_date and entry.due_date.date().strftime('%d/%m/%Y'),
                            f"Sort le {entry.release_date and entry.release_date.date().strftime('%d/%m/%Y')}" if entry.release_date else ""
                        ]
                    )
                k=0
                for k, row in enumerate(main_data, start=next_row + 1):
                    for j, value in enumerate(row, start=1):
                        cell = ws.cell(row=k, column=j, value=value)
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border
                next_row = k + 1


            # Ajustement largeur colonnes
            for col in ws.iter_cols(min_row=1, max_row=ws.max_row):
                col_letter = get_column_letter(col[0].column)
                max_length = 0
                for cell in col:
                    try:
                        # Ignorer la cellule G6
                        if cell.coordinate in cells_to_ignore_width:
                            continue
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except AttributeError:
                        continue  # Ignore merged cells
                ws.column_dimensions[col_letter].width = max_length + 2

        # Enregistrer en mémoire
        output = BytesIO()
        wb.save(output)
        file_data = output.getvalue()
        output.close()

        encoded_file = base64.b64encode(file_data).decode("utf-8")

        return encoded_file
    except Exception as e:
        raise e
